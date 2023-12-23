import copy
import glob
import os
import numpy as np
import openpyxl
class Process:
    def __init__(self, file_name):
        self.file_name = file_name
        self.array = self.read_val()[0]
        self.header = self.read_val()[1]
        self.all_array = self.read_val()[2]

    def read_val(self):
        data_list = []
        header = []
        all_arr = []
        data = open(self.file_name)
        for line in data:
            line = line.strip()
            line = line.split(",")
            all_arr.append(line)
            if line[0] == "OBJECTID":
                header.append(line)
                continue
            data_list.append(line)
        data.close()
        return data_list, header, all_arr

    def save_val(self, input_list,mode):
        name_cut_num = self.file_name.find("\\")
        name_cut_num = self.file_name[name_cut_num:].find("_") + name_cut_num + 1
        save_path = self.file_name[name_cut_num:-4]
        name_cut_num = save_path.rfind("_")
        save_path = save_path[:name_cut_num]
        name_cut_num = self.file_name.find("\\")
        save_path = self.file_name[:name_cut_num] + "\\" + save_path

        if mode == "npy":
            np.save(save_path+".npy", input_list)

        elif mode == "xlsx":
            wb = openpyxl.Workbook()
            sheet = wb.active
            rows_num = 0
            for cols in self.header[0]:
                sheet[chr(65+rows_num)+str(1)] = cols
                for rows in range(len(input_list)):
                    sheet[chr(65+rows_num)+str(rows+2)] = input_list[rows][rows_num]
                rows_num += 1
            wb.save(save_path+".xlsx")
            return save_path

    @staticmethod
    def sort_date(input_list):
        model = input_list[0]
        if len(model[1].split("/")[2]) == 4:
            sort_mode = [[1,1],[1,0],[1,2],[2,1],[2,0]]
        else:
            sort_mode = [[1,2],[1,1],[1,0],[2,1],[2,0]]

        sorted_list = copy.deepcopy(input_list)

        #处理站点数据异常
        error_num = []
        for i in range(len(input_list)):
            if "A" in sorted_list[i][2].split("-")[1] or "B" in sorted_list[i][2].split("-")[1]:
                sorted_list[i][2] = sorted_list[i][2].split("-")[0]+"-"+sorted_list[i][2].split("-")[1][0]+"0"
                error_n = int(sorted_list[i][0])
                error_num.append(error_n)

        #排序
        for i in range(3):
            sorted_list = sorted(sorted_list, key = lambda x: int(str(x[sort_mode[i][0]].split("/")[sort_mode[i][1]])) )
        for i in range(3,5):

            sorted_list = sorted(sorted_list, key = lambda x: int(str(x[sort_mode[i][0]].split("-")[sort_mode[i][1]])) )

        #返回异常站点原始值
        for i in error_num:
            for j in range(len(sorted_list)):
                if i == int(sorted_list[j][0]):
                    sorted_list[j][2] = input_list[i-1][2]
        return sorted_list

    @staticmethod
    def cut_excel(file_name, a, b, header):
        workbook = openpyxl.load_workbook(file_name)
        sheets = workbook.sheetnames
        sheet = workbook[sheets[0]]

        cut_list = []
        for i in sheet.iter_rows():
            col_list = []
            for cell in i:
                value = cell.value
                col_list.append(value)
            cut_list.append(col_list)
        cut_list = cut_list[a-1:b]

        wb = openpyxl.Workbook()
        sheet = wb.active
        rows_num = 0

        if header:
            for cols in data.header[0]:
                sheet[chr(65 + rows_num) + str(1)] = cols
                for rows in range(len(cut_list)):
                    sheet[chr(65 + rows_num) + str(rows + 2)] = cut_list[rows][rows_num]
                rows_num += 1
        else:
            for cols in cut_list[0]:
                for rows in range(len(cut_list)):
                    sheet[chr(65 + rows_num) + str(rows + 1)] = cut_list[rows][rows_num]
                rows_num += 1

        wb.save(file_name[:-4] +"_50.xlsx")

if __name__ == "__main__":
    file_path = glob.glob(r"Python_Programming&RS_Image_Processing_1\*.txt")
    for file in file_path:
        data = Process(file)
        data_sorted = data.sort_date(data.array)
        xlsx_save_path = data.save_val(data_sorted, "xlsx")
        data.save_val(data.all_array, "npy")
        Process.cut_excel(xlsx_save_path+".xlsx", 101, 150, data.header)


